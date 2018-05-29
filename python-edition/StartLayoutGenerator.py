from openpyxl import load_workbook
from lxml import etree
wb = load_workbook("layout.xlsx")

groups_range = wb['Groups']
folders_range = wb['Folders']
taskbar_range = wb['Taskbar']

xmlns_ns = "http://schemas.microsoft.com/Start/2014/LayoutModification"
defaultlayout_ns = "http://schemas.microsoft.com/Start/2014/FullDefaultLayout"
start_ns = "http://schemas.microsoft.com/Start/2014/StartLayout"
taskbar_ns = "http://schemas.microsoft.com/Start/2014/TaskbarLayout"

ns_map = {None: xmlns_ns, "defaultLayout": defaultlayout_ns, "start": start_ns, "taskbar": taskbar_ns}

root = etree.Element("LayoutModificationTemplate", Version="1", nsmap=ns_map)

layoutOptions = etree.SubElement(root, "LayoutOptions", StartTileGroupCellWidth="6")
defaultLayoutOverride = etree.SubElement(root, "DefaultLayoutOverride")
startLayoutCollection = etree.SubElement(defaultLayoutOverride, "StartLayoutCollection")
startLayout = etree.SubElement(startLayoutCollection, "{http://schemas.microsoft.com/Start/2014/FullDefaultLayout}StartLayout", GroupCellWidth="6")

#for each group until empty
for group in range (1, 16):
    if str(groups_range.cell(row=group, column=1).value) == "None":
        break
    currentGroup = etree.SubElement(startLayout, "{http://schemas.microsoft.com/Start/2014/StartLayout}Group", Name=str(groups_range["A"+str(group)].value))
    #for each element in group until empty
    for gelement in range (2, 16):
        if str(groups_range.cell(row=group, column=gelement).value) == "None":
            break
        #folder
        if str(groups_range.cell(row=group, column=gelement).value)[:6] == "Folder":
            foldernum = int(groups_range.cell(row=group, column=gelement).value[-1:])
            currentFolder = etree.SubElement(currentGroup, "{http://schemas.microsoft.com/Start/2014/StartLayout}Folder", Name=str(folders_range.cell(row=foldernum, column=1).value), Size="2x2", Column=str((gelement - 2)%3 * 2), Row=str(int((gelement - 2)/6 * 2)))
            for felement in range (2, 16):
                if str(folders_range.cell(row=foldernum, column=felement).value) == "None":
                    break
                etree.SubElement(currentFolder, "{http://schemas.microsoft.com/Start/2014/StartLayout}DesktopApplicationTile", Size="2x2", Column=str((felement - 2)%3 * 2), Row=str(int((felement - 2)/6 * 2)), DesktopApplicationID=str(folders_range.cell(row=foldernum, column=felement).value))
        #Not a folder
        else:
            etree.SubElement(currentGroup, "{http://schemas.microsoft.com/Start/2014/StartLayout}DesktopApplicationTile", Size="2x2", Column=str((gelement - 2)%3 * 2), Row=str(int((gelement - 2)/6 * 2)), DesktopApplicationID=str(groups_range.cell(row=group, column=gelement).value))

#Deal with Taskbar
if str(taskbar_range.cell(row=1, column=1).value) != "None":
    customTaskbarLayoutCollection = etree.SubElement(root, "CustomTaskbarLayoutCollection", PinListPlacement="Replace")
    taskbarLayout = etree.SubElement(customTaskbarLayoutCollection, "{http://schemas.microsoft.com/Start/2014/FullDefaultLayout}TaskbarLayout")
    taskbarPinList = etree.SubElement(taskbarLayout, "{http://schemas.microsoft.com/Start/2014/TaskbarLayout}TaskbarPinList")
    for telement in range (1, 16):
        if str(taskbar_range.cell(row=1, column=telement).value) == "None":
            break
        etree.SubElement(taskbarPinList, "{http://schemas.microsoft.com/Start/2014/TaskbarLayout}DesktopApp", DesktopApplicationLinkPath=str(taskbar_range.cell(row=1, column=telement).value))
        
        
s = etree.tostring(root, pretty_print=True)

et = etree.ElementTree(root)
et.write("StartLayout.xml", pretty_print=True)

#print (s)
